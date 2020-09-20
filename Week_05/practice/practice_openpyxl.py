import openpyxl
from openpyxl.chart import PieChart, Reference

wb = openpyxl.Workbook()  # 创建xlsx文件类
cs1 = wb.create_sheet("test1", 0)  # 表明+ index创建sheet
cs3 = wb.create_sheet("test3")  # 表名 创建sheet
cs4 = wb.create_sheet("test4")
cs5 = wb.create_sheet(index=10)  # index 创建sheet index大于当前标数量的时候加载表末尾
cs2 = wb.create_sheet("test2", 2)  # 插入 sheet

cs1['a1'] = "a1"  # 添加单个数据 字符串
cs1["A2"] = 2   # 添加单个数据 int
nums = [1, 2, 3, 4]
cs2.append(nums)   # add a line
cs2.append(nums*2)
cs2.append([i * 2 for i in nums])
#循环加入
matrix = [[1,2,3],[1,2,3],[1,2,3]]
i = 1
for row in matrix:
    i += 1
    cs4.append(row * i)
# cs3["a"] = nums * 5  #  AttributeError: 'tuple' object has no attribute 'value'
itemList = cs2['a']  # [(<Cell 'test2'.A1>, <Cell 'test2'.A2>, <Cell 'test2'.A3>)]
itemList2 = cs2["1"]
print('[itemList]', [itemList])
print('[itemList].value', [item.value for item in itemList])    # 按列取值
print('[itemList2]', [itemList2])   # 按行取值
print('[itemList2].value', [item.value for item in itemList2])
itemMatrix = [[item.value for item in row] for row in cs4["a:c"]]  # 多列取值
itemMatrix1 = [item.value for row in cs4["1:3"] for item in row]  # 多行取值
itemMatrix2 = [item.value for row in cs4["a1:c4"] for item in row]  # 范围取值

print(itemMatrix)
print(itemMatrix1)
print(itemMatrix2)
pie = PieChart()
label = Reference(cs4, 1, 2, 5)
data = Reference(cs4, 2, 2, 5)
pie.add_data(data)
pie.set_categories(label)
cs4.add_chart(pie,"a10")

wb.save("test.xlsx")
